import pandas
import requests
from openpyxl import Workbook
import json


result = requests.get("https://www.balldontlie.io/api/v1/players")
#print(result.text)
#df = pd.read_json("result.json")
#bn=pd.DataFrame(df.features.values.tolist())['id']
#pd.DataFrame.from_records(bn)
result_in_json = result.json()
#print(result_in_json)
df = pandas.json_normalize(result_in_json['data'])
#print(df)
#df.to_csv("./sujant.csv")

#This is for the convertion of CSV to HTML
#df = pd.read_csv("jsonoutput.csv")

#df.to_html("test2.htm", index= False)

#html_file = df.to_html()

# This is for the file to convert csv to xml
#wb = Workbook()
#ws = wb.active
#with open('jsonoutput.csv', 'r') as f:
 #   for row in csv.reader(f):
  #      ws.append(row)
#wb.save('csvtoxl.xlsx')

# Loading our Excel file
#wb = load_workbook("demo_database.xlsx")

# creating the sheet 1 object
#ws = wb.worksheets[0]

# Iterating rows for getting the values of each row
#for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
 #   print([cell.value for cell in row])
wb = Workbook()
ws = wb.active
with open('jsonoutput.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('csvtoxl.xlsx')